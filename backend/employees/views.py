import io
from django.db import transaction
from django.http import HttpResponse
from rest_framework import generics, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.utils import timezone
from django.core.cache import cache
from core.permissions import CanViewEmployees, CanEditEmployees, IsHRM, IsHR
from .models import Department, Site, Employee, AcademicQualification, EmployeeStatusLog
from .serializers import (
    DepartmentSerializer, SiteSerializer, EmployeeSerializer,
    EmployeeListSerializer, AcademicQualificationSerializer,
    EmployeeStatusLogSerializer
)
from . import bulk_import
from payroll.models import Payroll

# ── Cache helpers ─────────────────────────────────────────────────────────────
DEPT_LIST_KEY    = 'departments:list'
SITE_LIST_KEY    = 'sites:list'
EMPLOYEE_LIST_KEY = 'employees:list'   # role-scoped keys built at request time
CACHE_TTL        = 300  # 5 minutes


class DepartmentListCreateView(generics.ListCreateAPIView):
    permission_classes = (IsAuthenticated,)
    serializer_class   = DepartmentSerializer

    def list(self, request, *args, **kwargs):
        cached = cache.get(DEPT_LIST_KEY)
        if cached is not None:
            return Response(cached)
        response = super().list(request, *args, **kwargs)
        cache.set(DEPT_LIST_KEY, response.data, CACHE_TTL)
        return response

    def perform_create(self, serializer):
        serializer.save()
        cache.delete(DEPT_LIST_KEY)

    def get_queryset(self):
        return Department.objects.all().order_by('name')


class DepartmentDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = (IsHRM,)
    serializer_class   = DepartmentSerializer
    queryset           = Department.objects.all()

    def perform_update(self, serializer):
        serializer.save()
        cache.delete(DEPT_LIST_KEY)

    def perform_destroy(self, instance):
        if self.request.user.role not in ('IT', 'HRM'):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Only the IT Manager or HR Manager can delete records.')
        instance.delete()
        cache.delete(DEPT_LIST_KEY)

    def destroy(self, request, *args, **kwargs):
        if request.user.role not in ('IT', 'HRM'):
            return Response(
                {'error': 'Only the IT Manager or HR Manager can delete records.'},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().destroy(request, *args, **kwargs)


class SiteListCreateView(generics.ListCreateAPIView):
    permission_classes = (IsAuthenticated,)
    serializer_class   = SiteSerializer

    def list(self, request, *args, **kwargs):
        cached = cache.get(SITE_LIST_KEY)
        if cached is not None:
            return Response(cached)
        response = super().list(request, *args, **kwargs)
        cache.set(SITE_LIST_KEY, response.data, CACHE_TTL)
        return response

    def perform_create(self, serializer):
        serializer.save()
        cache.delete(SITE_LIST_KEY)

    def get_queryset(self):
        return Site.objects.all().order_by('name')


class SiteDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = (IsHRM,)
    serializer_class   = SiteSerializer
    queryset           = Site.objects.all()

    def perform_update(self, serializer):
        serializer.save()
        cache.delete(SITE_LIST_KEY)

    def destroy(self, request, *args, **kwargs):
        if request.user.role not in ('IT', 'HRM'):
            return Response(
                {'error': 'Only the IT Manager or HR Manager can delete records.'},
                status=status.HTTP_403_FORBIDDEN
            )
        response = super().destroy(request, *args, **kwargs)
        cache.delete(SITE_LIST_KEY)
        return response


class EmployeeListCreateView(generics.ListCreateAPIView):
    permission_classes = (CanViewEmployees,)

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return EmployeeSerializer
        return EmployeeListSerializer

    def get_queryset(self):
        user = self.request.user
        if user.role == 'HOD':
            return Employee.objects.filter(department=user.department)
        return Employee.objects.all().order_by('last_name')

    def _cache_key(self):
        user = self.request.user
        # Scope key by role so HODs never see another department's cache
        if user.role == 'HOD':
            return f'{EMPLOYEE_LIST_KEY}:hod:{user.pk}'
        return f'{EMPLOYEE_LIST_KEY}:all'

    def list(self, request, *args, **kwargs):
        key = self._cache_key()
        cached = cache.get(key)
        if cached is not None:
            return Response(cached)
        response = super().list(request, *args, **kwargs)
        cache.set(key, response.data, CACHE_TTL)
        return response

    def create(self, request, *args, **kwargs):
        if request.user.role not in ('IT', 'HRM', 'HR'):
            return Response(
                {'error': 'You do not have permission to add employees.'},
                status=status.HTTP_403_FORBIDDEN
            )
        response = super().create(request, *args, **kwargs)
        # Bust all employee list caches on write
        cache.delete_pattern(f'{EMPLOYEE_LIST_KEY}:*')
        return response


class EmployeeDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = (CanViewEmployees,)
    serializer_class   = EmployeeSerializer

    def get_queryset(self):
        user = self.request.user
        if user.role == 'HOD':
            return Employee.objects.filter(department=user.department)
        return Employee.objects.all()

    def update(self, request, *args, **kwargs):
        if not request.user.can_edit:
            return Response(
                {'error': 'You do not have permission to edit employees.'},
                status=status.HTTP_403_FORBIDDEN
            )
        response = super().update(request, *args, **kwargs)
        cache.delete_pattern(f'{EMPLOYEE_LIST_KEY}:*')
        return response

    def destroy(self, request, *args, **kwargs):
        if request.user.role != 'IT':
            return Response(
                {'error': 'Only the IT Manager can delete records.'},
                status=status.HTTP_403_FORBIDDEN
            )
        response = super().destroy(request, *args, **kwargs)
        cache.delete_pattern(f'{EMPLOYEE_LIST_KEY}:*')
        return response


class EmployeeStatusChangeView(APIView):
    permission_classes = (CanEditEmployees,)

    def post(self, request, pk):
        try:
            employee   = Employee.objects.get(pk=pk)
            new_status = request.data.get('status')
            reason     = request.data.get('reason', '')

            if new_status not in dict(Employee.STATUS_CHOICES):
                return Response({'error': 'Invalid status.'}, status=status.HTTP_400_BAD_REQUEST)

            if new_status in ('dismissed', 'retired', 'suspended') and not reason:
                return Response({'error': 'A reason is required for this status.'}, status=status.HTTP_400_BAD_REQUEST)

            EmployeeStatusLog.objects.create(
                employee   = employee,
                old_status = employee.status,
                new_status = new_status,
                reason     = reason,
                changed_by = request.user.username,
            )

            employee.status            = new_status
            employee.status_reason     = reason
            employee.status_changed_at = timezone.now()
            employee.save()

            cache.delete_pattern(f'{EMPLOYEE_LIST_KEY}:*')
            return Response({'message': f'Status updated to {new_status}.'}, status=status.HTTP_200_OK)

        except Employee.DoesNotExist:
            return Response({'error': 'Employee not found.'}, status=status.HTTP_404_NOT_FOUND)


class AcademicQualificationView(generics.ListCreateAPIView):
    permission_classes = (CanViewEmployees,)
    serializer_class   = AcademicQualificationSerializer

    def get_queryset(self):
        return AcademicQualification.objects.filter(employee_id=self.kwargs['pk'])

    def perform_create(self, serializer):
        employee = Employee.objects.get(pk=self.kwargs['pk'])
        serializer.save(employee=employee)


class EmployeeBulkImportView(APIView):
    """
    Bulk-create employees from an uploaded .xlsx or .csv file.
    Accepts every Employee field EXCEPT profile_picture, cv, and
    highest_education_certificate — those stay per-employee, uploaded
    individually through the normal add/edit form.

    A "Bank Name" / "Account Number" (+ optional "Basic Salary") column is
    also supported and is saved onto that employee's Payroll record as their
    USD account — the bank name is fuzzy-matched against Zimbabwe's
    commercial bank list (see bulk_import.resolve_bank_name) so entries like
    "NMB" resolve to "NMB Bank", matching the Add/Edit Employee dropdowns.

    For .xlsx/.xlsm files with multiple sheet tabs, each tab's name is treated
    as a Site (e.g. a tab named "Bindura" assigns every employee on that tab
    to a "Bindura" site, auto-creating it if it doesn't already exist). Sheets
    with a generic name (Sheet1, Employees, Instructions, etc.) or that don't
    look like employee data are skipped for site-assignment purposes.
    """
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        if request.user.role not in ('IT', 'HRM', 'HR'):
            return Response(
                {'error': 'You do not have permission to import employees.'},
                status=status.HTTP_403_FORBIDDEN
            )

        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'No file was uploaded.'}, status=status.HTTP_400_BAD_REQUEST)

        name = file_obj.name or ''
        ext = name.rsplit('.', 1)[-1].lower() if '.' in name else ''
        if ext not in ('xlsx', 'xlsm', 'csv'):
            return Response(
                {'error': 'Unsupported file type. Please upload a .xlsx or .csv file.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            sheets = bulk_import.read_workbook_sheets(file_obj, ext)
        except Exception as e:
            return Response({'error': f'Could not read the file: {e}'}, status=status.HTTP_400_BAD_REQUEST)

        # Drop fully blank rows, then drop sheets that end up with no rows at all
        cleaned_sheets = []
        for sheet_name, rows in sheets:
            rows = [r for r in rows if any(c not in (None, '') for c in r)]
            if rows:
                cleaned_sheets.append((sheet_name, rows))

        if not cleaned_sheets:
            return Response({'error': 'The file appears to be empty.'}, status=status.HTTP_400_BAD_REQUEST)

        # ── Preview vs Confirm ───────────────────────────────────────────────
        # By default this is a dry-run: nothing is written to the database.
        # It validates every row and returns a statistics summary (rows that
        # would be added, per-department / per-site breakdown, and which
        # departments/sites are brand new) so the admin can review before
        # committing. Pass confirm=true to actually perform the import.
        commit = str(request.data.get('confirm', '')).strip().lower() in ('1', 'true', 'yes', 'on')

        departments_by_name = {d.name.strip().lower(): d for d in Department.objects.all()}
        sites_by_name        = {s.name.strip().lower(): s for s in Site.objects.all()}
        seen_numbers      = set(Employee.objects.values_list('employee_number', flat=True))
        seen_national_ids = set(x.upper() for x in Employee.objects.values_list('national_id', flat=True))
        created_sites       = set()
        created_departments = set()
        dept_counts = {}   # resolved department name -> count of valid rows
        site_counts = {}   # resolved site name -> count of valid rows
        sheets_ignored  = []   # sheets with no recognisable First/Last Name columns

        created_count = 0
        banking_count = 0
        skipped = []
        total_rows = 0

        for sheet_name, rows in cleaned_sheets:
            header_row, data_rows = rows[0], rows[1:]
            field_map = bulk_import.map_headers(header_row)
            is_flat_template = 'first_name' in field_map.values() and 'last_name' in field_map.values()

            legacy_rows = None
            if not is_flat_template:
                # Falls back to parsing the raw HR roster layout: title rows,
                # then a header row whose first column is literally "NAME"
                # (a single combined name column) rather than separate
                # First Name / Last Name columns.
                legacy_rows = bulk_import.parse_legacy_roster_sheet(rows)

            if not is_flat_template and legacy_rows is None:
                sheets_ignored.append(sheet_name or 'Unnamed sheet')
                continue

            # Only single-sheet CSVs have no tab name; for xlsx, a generically-named
            # tab (Sheet1, Employees, Instructions...) isn't treated as a site name.
            default_site_name = sheet_name if bulk_import.sheet_name_is_site_like(sheet_name) else None

            if is_flat_template:
                row_iter = []
                for idx, raw_row in enumerate(data_rows, start=2):  # row 1 is the header
                    if not any(c not in (None, '') for c in raw_row):
                        continue
                    row_data = {}
                    for col_idx, field in field_map.items():
                        if col_idx < len(raw_row):
                            row_data[field] = raw_row[col_idx]
                    row_iter.append((idx, row_data))
            else:
                row_iter = legacy_rows  # already a list of (excel_row_number, row_data)

            for idx, row_data in row_iter:
                total_rows += 1

                if not row_data.get('site') and default_site_name:
                    row_data['site'] = default_site_name

                display_name = f"{row_data.get('first_name') or ''} {row_data.get('last_name') or ''}".strip() or f"Row {idx}"

                reason = bulk_import.process_row(
                    row_data, departments_by_name, sites_by_name,
                    seen_numbers, seen_national_ids, created_sites,
                    created_departments, commit,
                )
                if reason:
                    skipped.append({'row': idx, 'sheet': sheet_name, 'name': display_name, 'reason': reason})
                    continue

                dept_name = row_data.pop('_department_name', None)
                site_name = row_data.pop('_site_name', None)
                bank_name_usd    = row_data.pop('_bank_name_usd', '') or ''
                bank_account_usd = row_data.pop('_bank_account_usd', '') or ''
                basic_salary     = row_data.pop('_basic_salary', None)
                has_banking = bool(bank_name_usd or bank_account_usd or basic_salary)

                if commit:
                    try:
                        with transaction.atomic():
                            serializer = EmployeeSerializer(data=row_data)
                            if not serializer.is_valid():
                                errs = serializer.errors
                                msg = '; '.join(
                                    f"{k}: {', '.join(str(x) for x in v)}" for k, v in errs.items()
                                )
                                raise ValueError(msg or 'Invalid data.')
                            employee = serializer.save()
                            if has_banking:
                                Payroll.objects.update_or_create(
                                    employee=employee,
                                    defaults={
                                        'basic_salary': basic_salary or 0,
                                        'bank_name_usd': bank_name_usd,
                                        'bank_account_usd': bank_account_usd,
                                        'currency': 'USD',
                                        'updated_by': request.user.username,
                                    },
                                )
                    except ValueError as ve:
                        skipped.append({'row': idx, 'sheet': sheet_name, 'name': display_name, 'reason': str(ve)})
                        continue
                    except Exception as e:
                        skipped.append({'row': idx, 'sheet': sheet_name, 'name': display_name, 'reason': f'Unexpected error: {e}'})
                        continue

                    seen_numbers.add(employee.employee_number)
                    seen_national_ids.add(employee.national_id.upper())
                    if has_banking:
                        banking_count += 1
                else:
                    # Preview / dry-run: validate only, don't write anything.
                    serializer = EmployeeSerializer(data=row_data)
                    if not serializer.is_valid():
                        errs = serializer.errors
                        msg = '; '.join(
                            f"{k}: {', '.join(str(x) for x in v)}" for k, v in errs.items()
                        )
                        skipped.append({'row': idx, 'sheet': sheet_name, 'name': display_name, 'reason': msg or 'Invalid data.'})
                        continue
                    # Still track it as "seen" for this request so within-file
                    # duplicate employee numbers/national IDs are caught.
                    seen_numbers.add(row_data.get('employee_number'))
                    seen_national_ids.add(str(row_data.get('national_id', '')).upper())
                    if has_banking:
                        banking_count += 1

                created_count += 1
                dept_key = dept_name or '— No Department —'
                dept_counts[dept_key] = dept_counts.get(dept_key, 0) + 1
                site_key = site_name or '— No Site —'
                site_counts[site_key] = site_counts.get(site_key, 0) + 1

        if commit and created_count:
            cache.delete_pattern(f'{EMPLOYEE_LIST_KEY}:*')
        if commit and created_sites:
            cache.delete(SITE_LIST_KEY)
        if commit and created_departments:
            cache.delete(DEPT_LIST_KEY)

        departments_summary = sorted(
            [{'name': n, 'count': c, 'is_new': n in created_departments} for n, c in dept_counts.items()],
            key=lambda x: (-x['count'], x['name']),
        )
        sites_summary = sorted(
            [{'name': n, 'count': c, 'is_new': n in created_sites} for n, c in site_counts.items()],
            key=lambda x: (-x['count'], x['name']),
        )

        return Response({
            'preview':              not commit,
            'total_rows':           total_rows,
            'created':              created_count,   # rows added (commit) / rows that would be added (preview)
            'valid_count':          created_count,
            'skipped_count':        len(skipped),
            'skipped':              skipped,
            'with_banking':         banking_count,
            'sites_created':        sorted(created_sites),
            'departments_created':  sorted(created_departments),
            'sheets_ignored':       sheets_ignored,
            'departments_summary':  departments_summary,
            'sites_summary':        sites_summary,
        }, status=status.HTTP_200_OK)


class EmployeeBulkImportTemplateView(APIView):
    """Serves a ready-to-fill .xlsx template with the correct column headers."""
    permission_classes = (IsAuthenticated,)

    COLUMNS = [
        'Employee Number', 'First Name', 'Last Name', 'Middle Name',
        'Date of Birth', 'National ID', 'Gender', 'Phone Number', 'Email', 'Address',
        'Department', 'Job Title', 'Date Joined', 'Employment Type', 'Status', 'Status Reason',
        'Contract Start', 'Contract End', 'Highest Education',
        'Next of Kin Name', 'Next of Kin Relationship', 'Next of Kin Phone',
        'Next of Kin Email', 'Next of Kin National ID', 'Next of Kin Address',
        # Saved onto the employee's Payroll record as their USD account. Bank
        # names are fuzzy-matched (e.g. "NMB" -> "NMB Bank") so they don't
        # need to match the dropdown text exactly.
        'Bank Name', 'Account Number', 'Basic Salary',
    ]

    EXAMPLE_ROW = [
        '', 'John', 'Moyo', '',
        '1990-05-14', '63-1234567A12', 'Male', '0771234567', 'john.moyo@example.com', '12 Samora Machel Ave, Harare',
        'Finance', 'Accountant', '2024-01-15', 'Full Time', 'Employed', '',
        '', '', '',
        'Jane Moyo', 'Spouse', '0779876543', '', '', '',
        'NMB Bank', '1002003004', '450',
    ]

    def get(self, request):
        if request.user.role not in ('IT', 'HRM', 'HR'):
            return Response(
                {'error': 'You do not have permission to download this template.'},
                status=status.HTTP_403_FORBIDDEN
            )

        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Employees'
        ws.append(self.COLUMNS)
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='1557B0', end_color='1557B0', fill_type='solid')
        ws.append(self.EXAMPLE_ROW)
        for i, _ in enumerate(self.COLUMNS):
            ws.column_dimensions[get_column_letter(i + 1)].width = 20

        notes = wb.create_sheet('Instructions')
        instructions = [
            ['Field', 'Notes'],
            ['Employee Number', 'Leave blank to auto-generate.'],
            ['Date of Birth / Date Joined / Contract dates', 'Use YYYY-MM-DD (or any common date format).'],
            ['National ID', 'Format: DD-NNNNNN(N)LNN, e.g. 63-207522S72'],
            ['Gender', 'Male, Female, or Other'],
            ['Department', 'If it matches an existing department name, it\'s assigned automatically. If not, the employee is still added — just without a department — rather than being skipped.'],
            ['Employment Type', 'Full Time, Part Time, or Contract (Contract requires Contract Start & End)'],
            ['Status', 'Employed, Retired, Dismissed, Resigned, or Suspended (defaults to Employed)'],
            ['Highest Education', 'O Level, A Level, Certificate, Diploma, Degree, Honours, Masters, or PhD'],
            ['Sites (multiple locations)', 'Put each site\'s employees on their own sheet tab, and name the tab after the site (e.g. "Bindura", "Bluffhill"). The site is created automatically — don\'t use generic tab names like "Sheet1" if you want a site assigned.'],
            ['Bank Name / Account Number / Basic Salary', 'Optional — if given, saved as the employee\'s USD bank account on their payroll record. Bank Name is matched loosely against Zimbabwe\'s commercial banks (e.g. "NMB" becomes "NMB Bank"), so it doesn\'t need to be typed exactly. Leave Basic Salary blank to fill it in later.'],
            ['Profile picture / CV / certificates', 'Not supported via import — add these per-employee afterwards.'],
        ]
        for row in instructions:
            notes.append(row)
        for cell in notes[1]:
            cell.font = Font(bold=True)
        notes.column_dimensions['A'].width = 40
        notes.column_dimensions['B'].width = 70

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="employee_import_template.xlsx"'
        return response


class EmployeeExportBySiteView(APIView):
    """
    Exports all employees as an .xlsx workbook with one sheet per Site,
    matching the company's "Employee Information Sheet" layout: a bold
    site-name title, a header row, and employees grouped under bold
    department-name section headers. Employees with no site are grouped
    onto an "Unassigned" sheet.
    """
    permission_classes = (IsAuthenticated, CanViewEmployees)

    COLUMNS = [
        'NAME', 'POSITION', 'DEPARTMENT', 'Start Date', 'ID/Passport Number',
        'Date of Birth', 'Gender', 'Marital Status', 'Nationality', 'Home Add',
        'Contact phone No', 'Emergency Contact Number', 'Name', 'Relationship',
        'Bank Name', 'Account Number', 'Spouse Name', 'Contact Number',
        'ID Number', 'Names of Children and their IDs',
    ]

    GENDER_LABELS = {'M': 'Male', 'F': 'Female', 'O': 'Other'}

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        employees = (
            Employee.objects
            .select_related('site', 'department')
            .prefetch_related('payroll')
            .order_by('site__name', 'department__name', 'first_name', 'last_name')
        )

        # Optional filter: ?pay_type=daily | monthly | all (default: all)
        pay_type = (request.query_params.get('pay_type') or 'all').lower()
        if pay_type in ('daily', 'monthly'):
            employees = [
                emp for emp in employees
                if getattr(getattr(emp, 'payroll', None), 'pay_type', 'monthly') == pay_type
            ]

        by_site = {}
        for emp in employees:
            site_name = emp.site.name if emp.site else 'Unassigned'
            by_site.setdefault(site_name, []).append(emp)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        if not by_site:
            ws = wb.create_sheet('No Employees')
            label = {'daily': 'daily-rate', 'monthly': 'monthly-salary'}.get(pay_type, '')
            ws['A1'] = f'No {label} employees found.'.replace('  ', ' ')
            ws.column_dimensions['A'].width = 40
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            response = HttpResponse(
                buf.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename_suffix = {'daily': 'daily_rate', 'monthly': 'monthly_salary'}.get(pay_type, 'all')
            response['Content-Disposition'] = f'attachment; filename="employees_by_site_{filename_suffix}.xlsx"'
            return response

        title_font   = Font(name='Times New Roman', bold=True, size=11, underline='single')
        header_font  = Font(name='Times New Roman', bold=True, size=11)
        dept_font    = Font(name='Times New Roman', bold=True, size=11)
        dept_fill    = PatternFill('solid', fgColor='D9E2F3')
        body_font    = Font(name='Times New Roman', size=11)

        used_names = set()
        for site_name in sorted(by_site.keys()):
            safe_name = site_name[:31] or 'Site'
            base, n = safe_name, 1
            while safe_name in used_names:
                n += 1
                safe_name = f"{base[:28]}({n})"
            used_names.add(safe_name)

            ws = wb.create_sheet(safe_name)
            ws['A1'] = site_name
            ws['A1'].font = title_font

            ws.append([])
            ws.append(self.COLUMNS)
            for cell in ws[3]:
                cell.font = header_font

            emps = by_site[site_name]
            by_dept = {}
            for emp in emps:
                dept_name = emp.department.name if emp.department else 'Unassigned'
                by_dept.setdefault(dept_name, []).append(emp)

            for dept_name in sorted(by_dept.keys()):
                header_row = ws.max_row + 1
                ws.cell(row=header_row, column=1, value=dept_name)
                for col in range(1, len(self.COLUMNS) + 1):
                    ws.cell(row=header_row, column=col).fill = dept_fill
                ws.cell(row=header_row, column=1).font = dept_font

                for emp in by_dept[dept_name]:
                    payroll = getattr(emp, 'payroll', None)
                    full_name = ' '.join(p for p in [emp.first_name, emp.middle_name, emp.last_name] if p)
                    row = [
                        full_name,
                        emp.job_title,
                        dept_name,
                        emp.date_joined.strftime('%Y-%m-%d') if emp.date_joined else '',
                        emp.national_id,
                        emp.date_of_birth.strftime('%Y-%m-%d') if emp.date_of_birth else '',
                        self.GENDER_LABELS.get(emp.gender, emp.gender),
                        '',  # Marital Status — not tracked in HRPortal
                        '',  # Nationality — not tracked in HRPortal
                        emp.address,
                        emp.phone_number,
                        emp.nok_phone,
                        emp.nok_full_name,
                        emp.nok_relationship,
                        payroll.bank_name_usd if payroll else '',
                        payroll.bank_account_usd if payroll else '',
                        '',  # Spouse Name — not tracked in HRPortal
                        '',  # Spouse Contact Number — not tracked in HRPortal
                        '',  # Spouse ID Number — not tracked in HRPortal
                        '',  # Names of Children and their IDs — not tracked in HRPortal
                    ]
                    r = ws.max_row + 1
                    for col, val in enumerate(row, start=1):
                        c = ws.cell(row=r, column=col, value=val)
                        c.font = body_font

                ws.append([])

            widths = [30, 32, 20, 12, 20, 14, 10, 14, 14, 40, 16, 20, 20, 14, 16, 18, 16, 16, 14, 32]
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename_suffix = {'daily': 'daily_rate', 'monthly': 'monthly_salary'}.get(pay_type, 'all')
        response['Content-Disposition'] = f'attachment; filename="employees_by_site_{filename_suffix}.xlsx"'
        return response